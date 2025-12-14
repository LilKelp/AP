"""
Generate payment-list workbooks for each region under 02-inputs/Payment run raw/<REGION>
from SAP ALV exports (.xlsx or .xls).

Vendor lookups default to the OneDrive workbook
`OneDrive - novabio.onmicrosoft.com/Desktop/AZ Working Notes.xlsx`
(AU AP sheet cols W:X, NZ AP sheet cols U:V). If unavailable, fall back to the
local vendor files under 02-inputs/Payment run raw/.

Steps for each raw workbook:
1. Load the matching vendor list to map Vendor IDs to supplier names.
2. Ensure Sheet1 contains all raw records plus a SUPPLIER NAME column.
3. Add Sheet2 with a PivotTable laid out as Supplier -> Vendor -> DD -> Reference
   so overdue items can be filtered directly via the DD field. Supplier totals remain.

Usage:
    python 01-system/tools/ops/payment-list/payment_routine.py
"""

from __future__ import annotations

import ctypes
import re
import sys
from pathlib import Path
import tempfile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import win32com.client as win32

BASE_DIR = Path(__file__).resolve().parents[4]
ONEDRIVE_VENDOR_PATH = (
    Path.home()
    / "OneDrive - novabio.onmicrosoft.com"
    / "Desktop"
    / "AZ Working Notes.xlsx"
)
INPUT_ROOT = BASE_DIR / "02-inputs" / "Payment run raw"
OUTPUT_ROOT = BASE_DIR / "03-outputs" / "payment-list"

REGIONS = [
    {
        "code": "AU",
        "data_dir": INPUT_ROOT / "AU",
        "vendor_sources": [
            {
                "path": ONEDRIVE_VENDOR_PATH,
                "sheet": "AU AP",
                "usecols": "W:X",
                "copy_for_read": True,
            },
            {
                "path": INPUT_ROOT / "AU Vendor list.xlsx",
                "sheet": "Sheet3",
                "usecols": [0, 1],
            },
        ],
    },
    {
        "code": "NZ",
        "data_dir": INPUT_ROOT / "NZ",
        "vendor_sources": [
            {
                "path": ONEDRIVE_VENDOR_PATH,
                "sheet": "NZ AP",
                "usecols": "U:V",
                "copy_for_read": True,
            },
            {
                "path": INPUT_ROOT / "NZ Vendor list.xlsx",
                "sheet": "Sheet3",
                "usecols": [0, 1],
            },
        ],
    },
]


REQUIRED_COLUMNS = {
    "Vendor",
    "Reference",
    "DD",
    "Amount in local cur.",
}


def normalize_header_cell(value: object) -> str:
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text.lower()


def find_header_row(preview: pd.DataFrame) -> int:
    """Find the first row that looks like the SAP export header."""
    header_synonyms = {
        "vendor": {"vendor"},
        "reference": {"reference", "inv. ref.", "inv ref"},
        "dd": {"dd", "net due dt", "due date"},
        "amount": {
            "amount in local cur.",
            "amount in local cur",
            "amount in local currency",
            "lc amnt",
        },
    }
    for idx, row in preview.iterrows():
        cells = {
            normalize_header_cell(v)
            for v in row.values.tolist()
            if pd.notna(v)
        }
        if "vendor" not in cells:
            continue
        matches = sum(1 for synonyms in header_synonyms.values() if cells & synonyms)
        if matches >= 2:
            return int(idx)
    return 0


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Strip/rename SAP columns to the canonical names used by the pivot."""
    cleaned = []
    for col in df.columns:
        name = str(col).strip() if col is not None else ""
        name = re.sub(r"\s+", " ", name)
        cleaned.append(name)
    df.columns = cleaned

    rename_map: dict[str, str] = {}
    for col in df.columns:
        key = normalize_header_cell(col)
        if key in {"dd", "net due dt", "due date"}:
            rename_map[col] = "DD"
        elif key in {
            "amount in local cur.",
            "amount in local cur",
            "amount in local currency",
            "lc amnt",
        }:
            rename_map[col] = "Amount in local cur."
        elif key in {"inv. ref.", "inv ref"} and "Reference" not in df.columns:
            rename_map[col] = "Reference"
        elif key == "vendor":
            rename_map[col] = "Vendor"
        elif key == "reference":
            rename_map[col] = "Reference"

    df = df.rename(columns=rename_map)
    df = df.loc[
        :,
        [
            c
            for c in df.columns
            if c and not normalize_header_cell(c).startswith("unnamed")
        ],
    ]
    df = coalesce_duplicate_columns(
        df,
        "DD",
        invalid_values={"dd", " dd", "DD", " DD", ""},
    )
    df = coalesce_duplicate_columns(df, "Reference")
    df = coalesce_duplicate_columns(df, "Amount in local cur.")
    return df


def coalesce_duplicate_columns(
    df: pd.DataFrame, base_name: str, invalid_values: set[str] | None = None
) -> pd.DataFrame:
    """Collapse duplicate columns like DD/DD.1 into a single base_name column."""
    positions = [
        i
        for i, col in enumerate(df.columns)
        if col == base_name or str(col).startswith(f"{base_name}.")
    ]
    if len(positions) <= 1:
        return df

    best_pos = positions[0]
    best_score = -1
    for pos in positions:
        series = df.iloc[:, pos]
        score = int(series.notna().sum())
        if invalid_values:
            lowered = series.astype(str).str.lower().str.strip()
            score -= int(lowered.isin(invalid_values).sum())
        if score > best_score:
            best_score = score
            best_pos = pos

    best_series = df.iloc[:, best_pos]
    insert_at = min(positions)
    df = df.drop(df.columns[positions], axis=1)
    df.insert(insert_at, base_name, best_series)
    return df


def looks_like_ascii_export(preview: pd.DataFrame) -> bool:
    """Detect SAP 'text list saved as .xls' exports (single column with | separators)."""
    if preview.shape[1] != 1:
        return False
    col = preview.iloc[:, 0].dropna().astype(str)
    return bool(col.str.contains(r"\|").any())


def parse_ascii_export(lines: list[str]) -> pd.DataFrame:
    """Parse a SAP text list export into a structured DataFrame."""
    header_idx = None
    for i, line in enumerate(lines):
        lower = line.lower()
        if "|" in line and "vendor" in lower and "reference" in lower:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Could not find header row in text-list export.")

    header_line = lines[header_idx].strip()
    columns = [c.strip() for c in header_line.strip("|").split("|")]
    columns = [c for c in columns if c]

    rows: list[list[str]] = []
    for line in lines[header_idx + 1 :]:
        stripped = line.strip()
        if not stripped.startswith("|"):
            continue
        if "vendor" in stripped.lower() and "reference" in stripped.lower():
            continue
        if set(stripped) <= {"-", "|"}:
            continue
        parts = [p.strip() for p in stripped.strip("|").split("|")]
        if len(parts) < len(columns):
            parts += [""] * (len(columns) - len(parts))
        if len(parts) > len(columns):
            parts = parts[: len(columns)]
        rows.append(parts)

    if not rows:
        raise ValueError("No data rows found in text-list export.")
    return pd.DataFrame(rows, columns=columns)


def parse_amount_series(series: pd.Series) -> pd.Series:
    """Convert SAP amount strings like '341,199.00-' to floats."""
    def to_number(val):
        if pd.isna(val):
            return None
        if isinstance(val, (int, float)):
            return float(val)
        text = str(val).strip().replace(",", "").replace(" ", "")
        if not text:
            return None
        negative = text.endswith("-")
        if negative:
            text = text[:-1]
        try:
            num = float(text)
        except ValueError:
            return None
        return -num if negative else num

    return series.apply(to_number)


def load_raw_dataframe(data_path: Path) -> pd.DataFrame:
    """Load a SAP export (.xlsx or .xls) with header/column normalization."""
    temp_path: Path | None = None
    readable_path = data_path
    if data_path.suffix.lower() == ".xls":
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(data_path))
        try:
            tmp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            temp_path = Path(tmp_file.name)
            tmp_file.close()
            wb.SaveAs(str(temp_path), FileFormat=51)
            readable_path = temp_path
        finally:
            wb.Close(SaveChanges=False)
            excel.Quit()

    try:
        preview = pd.read_excel(readable_path, header=None, nrows=200)
        if looks_like_ascii_export(preview):
            raw_lines = (
                pd.read_excel(readable_path, header=None)
                .iloc[:, 0]
                .dropna()
                .astype(str)
                .tolist()
            )
            df = parse_ascii_export(raw_lines)
        else:
            header_row = find_header_row(preview)
            df = pd.read_excel(readable_path, header=header_row)
    finally:
        if temp_path and temp_path.exists():
            temp_path.unlink(missing_ok=True)

    df = normalize_columns(df)
    df = df.dropna(how="all")
    if "Vendor" in df.columns:
        df = df[df["Vendor"].notna()]

    missing = REQUIRED_COLUMNS - set(df.columns)
    if missing:
        raise ValueError(
            "Raw export is missing required columns: "
            + ", ".join(sorted(missing))
            + ". Please include these fields in the SAP export."
        )
    df["Amount in local cur."] = parse_amount_series(df["Amount in local cur."])
    df["DD"] = pd.to_datetime(df["DD"], dayfirst=True, errors="coerce")
    return df


def copy_with_winapi(src: Path, dst: Path) -> None:
    """Use Windows API copy to avoid share/lock issues when reading vendor workbooks."""
    result = ctypes.windll.kernel32.CopyFileW(str(src), str(dst), False)
    if result == 0:
        raise ctypes.WinError()


def load_vendor_lookup(vendor_sources: list[dict]) -> dict[int, str]:
    """Return {vendor_id: supplier_name} using the first available vendor source."""
    last_error: Exception | None = None
    for source in vendor_sources:
        path = Path(source["path"])
        sheet = source.get("sheet")
        usecols = source.get("usecols")
        copy_for_read = bool(source.get("copy_for_read"))
        if not path.exists():
            print(f"[WARN] Vendor source missing: {path}")
            continue

        temp_path: Path | None = None
        target_path = path
        try:
            if copy_for_read:
                temp_file = tempfile.NamedTemporaryFile(
                    suffix=path.suffix, delete=False
                )
                temp_path = Path(temp_file.name)
                temp_file.close()
                copy_with_winapi(path, temp_path)
                target_path = temp_path

            df = pd.read_excel(target_path, sheet_name=sheet, usecols=usecols)
            df = df.dropna()
            lookup: dict[int, str] = {}
            for _, row in df.iterrows():
                try:
                    vendor_id = int(row.iloc[0])
                except (TypeError, ValueError):
                    continue
                name = str(row.iloc[1]).strip()
                if name:
                    lookup[vendor_id] = name
            if lookup:
                print(f"[INFO] Vendor source loaded: {path}")
                return lookup
        except Exception as exc:  # pragma: no cover - defensive logging
            last_error = exc
            print(f"[WARN] Failed vendor source {path}: {exc}")
        finally:
            if temp_path and temp_path.exists():
                temp_path.unlink(missing_ok=True)

    if last_error:
        raise last_error
    raise FileNotFoundError(
        f"No vendor source available from: {[s['path'] for s in vendor_sources]}"
    )


def resolve_supplier(vendor_value, current_name, lookup: dict[int, str]) -> str:
    """Return the supplier name from the lookup, falling back to existing text."""
    if isinstance(current_name, str) and current_name.strip():
        return current_name.strip()
    if pd.isna(vendor_value):
        return "Unknown Supplier"
    try:
        vendor_id = int(vendor_value)
    except (TypeError, ValueError):
        return str(vendor_value)
    return lookup.get(vendor_id, f"Vendor {vendor_id}")


def ensure_supplier_column(df: pd.DataFrame, lookup: dict[int, str]) -> pd.DataFrame:
    """Add/populate the SUPPLIER NAME column next to Vendor."""
    if "SUPPLIER NAME" not in df.columns:
        vendor_idx = df.columns.get_loc("Vendor") if "Vendor" in df.columns else -1
        insert_at = vendor_idx + 1 if vendor_idx >= 0 else len(df.columns)
        df.insert(insert_at, "SUPPLIER NAME", None)

    df["SUPPLIER NAME"] = [
        resolve_supplier(vendor, current_name, lookup)
        for vendor, current_name in zip(df["Vendor"], df["SUPPLIER NAME"])
    ]
    return df


def write_base_workbook(df: pd.DataFrame, output_path: Path) -> tuple[int, int]:
    """Write Sheet1 with raw data + supplier names; return (row_count, col_count)."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Sheet1", index=False)

    wb = load_workbook(output_path)
    for sheet_name in list(wb.sheetnames):
        if sheet_name != "Sheet1":
            del wb[sheet_name]

    ws_summary = wb.create_sheet("Sheet2")
    ws_summary["A1"] = "Payment pivot (DD visible in rows for manual screening)"
    ws_summary[
        "A2"
    ] = "Filter DD entries or collapse totals to focus on overdue vs not due items."
    wb.save(output_path)
    return len(df.index) + 1, len(df.columns)


def add_pivot_table(output_path: Path, last_row: int, last_col: int) -> None:
    """Create the Excel pivot table on Sheet2 using COM automation."""
    last_col_letter = get_column_letter(last_col)
    source_range = f"Sheet1!A1:{last_col_letter}{last_row}"

    xl_database = 1
    xl_row_field = 1
    xl_sum = -4157
    xl_tabular_row = 1
    xl_pivot_version = 6

    excel = win32.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    workbook = excel.Workbooks.Open(str(output_path))
    try:
        ws_pivot = workbook.Worksheets("Sheet2")
        ws_pivot.Cells.Clear()
        ws_pivot.Range("A1").Value = (
            "Payment pivot (DD visible in rows for manual screening)"
        )
        ws_pivot.Range("A2").Value = (
            "Filter DD entries or collapse totals to focus on overdue items."
        )

        pivot_cache = workbook.PivotCaches().Create(
            SourceType=xl_database,
            SourceData=source_range,
            Version=xl_pivot_version,
        )
        pivot_table = pivot_cache.CreatePivotTable(
            TableDestination=ws_pivot.Range("A4"), TableName="PaymentPivot"
        )

        supplier_field = pivot_table.PivotFields("SUPPLIER NAME")
        supplier_field.Orientation = xl_row_field
        supplier_field.Position = 1
        supplier_field.Subtotals = [True] + [False] * 11

        vendor_field = pivot_table.PivotFields("Vendor")
        vendor_field.Orientation = xl_row_field
        vendor_field.Position = 2
        vendor_field.Subtotals = [False] * 12

        dd_field = pivot_table.PivotFields("DD")
        dd_field.Orientation = xl_row_field
        dd_field.Position = 3
        dd_field.Subtotals = [False] * 12

        reference_field = pivot_table.PivotFields("Reference")
        reference_field.Orientation = xl_row_field
        reference_field.Position = 4
        reference_field.Subtotals = [False] * 12

        data_field = pivot_table.AddDataField(
            pivot_table.PivotFields("Amount in local cur."),
            "Sum of Amount in local cur.",
            xl_sum,
        )
        data_field.NumberFormat = "#,##0.00"

        pivot_table.RowAxisLayout(xl_tabular_row)
    finally:
        workbook.Close(SaveChanges=True)
        excel.Quit()


def process_workbook(region_code: str, data_path: Path, lookup: dict[int, str]) -> Path:
    """Create the payment workbook for a single region/input file."""
    df = load_raw_dataframe(data_path)
    df = ensure_supplier_column(df, lookup)

    output_path = (
        OUTPUT_ROOT
        / region_code
        / f"PMT_{region_code}_{data_path.stem}.xlsx"
    )
    last_row, last_col = write_base_workbook(df, output_path)
    add_pivot_table(output_path, last_row, last_col)
    return output_path


def process_region(region_config: dict[str, object]) -> list[Path]:
    """Process all XLSX files for a region; return list of generated paths."""
    region_code = region_config["code"]
    data_dir = region_config["data_dir"]
    vendor_sources = region_config["vendor_sources"]

    if not data_dir.exists():
        print(f"[WARN] Data directory missing for {region_code}: {data_dir}")
        return []

    lookup = load_vendor_lookup(vendor_sources)
    generated_paths: list[Path] = []
    workbooks = [
        *data_dir.glob("*.xlsx"),
        *data_dir.glob("*.xls"),
    ]
    for workbook in sorted(
        [w for w in workbooks if not w.name.startswith("~$")]
    ):
        print(f"[INFO] Generating payment list for {region_code}: {workbook.name}")
        output_path = process_workbook(region_code, workbook, lookup)
        generated_paths.append(output_path)
    return generated_paths


def main() -> int:
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    all_outputs: list[Path] = []
    for region in REGIONS:
        outputs = process_region(region)
        all_outputs.extend(outputs)
    if not all_outputs:
        print("No payment workbooks were generated.")
        return 1
    print("\nCreated the following payment workbooks:")
    for path in all_outputs:
        print(f"  - {path.relative_to(BASE_DIR)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
